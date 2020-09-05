import os
import time

import yaml


# define template exception
class TemplateException(Exception):
    pass


# define template set value exception
class TemplateSetValueException(Exception):
    pass


# define source file exception
class SourceFileException(Exception):
    pass


# define extract value exception
class ExtractValueException(Exception):
    pass


# start extract
def extract():
    # read in the set of extract folders
    with open(os.getcwd() + '\\appconfig.yaml') as app_config_file:
        extract_folders = yaml.full_load(app_config_file)

        # loop and process each extract folder
        for extract_folder in extract_folders['extractFolders']:
            extract_set(extract_folder)

    app_config_file.close()


# process each extract set
def extract_set(extract_folder):
    config_filename = extract_folder + 'extractconfig.yaml'

    try:
        print('____________________\nPROCESSING ' + extract_folder)

        # read the extract set config
        with open(config_filename) as folder_config_file:
            extract_config = yaml.full_load(folder_config_file)
        folder_config_file.close()

        workbook = get_template_workbook(extract_folder, extract_config['templateFilename'])

    except TemplateException:
        print('ERROR reading template file in ' + config_filename)
    except Exception as ex:
        print('ERROR trying to process ' + config_filename)
        print(ex.args)
    else:
        # keep an error count
        error_count = 0

        # loop through and process each extract detail
        for extract_detail in extract_config['extractDetails']:
            try:
                # extract value
                value = extract_value(extract_folder, extract_detail)

                # set value to workbook
                set_value_to_template(workbook, extract_detail['templateSheet'], extract_detail['templateCell'], value)
            except SourceFileException:
                print('ERROR trying to read source file ' + extract_folder + extract_detail['sourceFilename'])
                error_count += 1
            except ExtractValueException:
                print('ERROR trying to extract value from')
                print(extract_detail)
                error_count += 1
            except TemplateSetValueException:
                print('ERROR trying to set value to template')
                print(extract_detail)
                error_count += 1

        if not error_count:
            # save the workbook
            write_template(workbook, extract_folder, extract_config['templateFilename'])
        else:
            print('ERROR count encountered: ' + str(error_count) + ', while processing ' + extract_folder)


# write template
def write_template(workbook, extract_folder, template_filename):
    # write as a new file with timestamp
    timestamp = time.strftime('%Y%m%d%H%M', time.localtime())
    filename = extract_folder + timestamp + '_' + template_filename

    workbook.save(filename)
    print('SUCCESSFULLY written out ' + filename)


# extract single value given folder and details
def extract_value(extract_folder, extract_detail):
    try:
        source_file = open(extract_folder + extract_detail['sourceFilename'])
    except Exception:
        raise SourceFileException()

    # read lines until key is found
    found = False
    while True:
        line = source_file.readline()

        if not line:
            break

        # check if line contains the key
        index = line.find(extract_detail['key'])
        if index != -1 and index == extract_detail['keyColStart'] - 1:
            found = True
            break

    if not found:
        raise ExtractValueException()
    else:
        # find value line by skipping offset
        for x in range(extract_detail['valueRowOffset']):
            line = source_file.readline()

        # line at this point should contain value
        source_file.close()
        value = line[extract_detail['valueColStart'] - 1:
                     extract_detail['valueColStart'] - 1 + extract_detail['valueColLength']]

        if value is None:
            raise ExtractValueException()
        else:
            print('Read value [' + value + '] from ')
            print(extract_detail)
            return float(value)


# get the workbook
def get_template_workbook(extract_folder, filename):
    from openpyxl import load_workbook
    try:
        workbook = load_workbook(extract_folder + filename)
        return workbook
    except Exception:
        raise TemplateException()


# write value to the template
def set_value_to_template(workbook, sheet_num, cell_address, value):
    try:
        worksheet = workbook[workbook.sheetnames[sheet_num - 1]]
        worksheet[cell_address] = value
    except Exception:
        raise TemplateSetValueException()


# start extract
extract()
